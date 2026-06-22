"""
Dashify — In-Memory Excel Parser
Ported from dash_no_ai.py parse_uploaded_excel() with security hardening.

Security:
- All processing in-memory via BytesIO (no disk writes)
- defusedxml patches openpyxl's XML reader to block XXE attacks
- Read-only workbook mode prevents formula evaluation
"""

import collections
import hashlib
import logging
import re
import threading
from io import BytesIO
from typing import Any

# Patch XML parsers BEFORE importing openpyxl to block XXE
import defusedxml.ElementTree
import xml.etree.ElementTree

# Monkey-patch the standard library XML parser with the safe version
xml.etree.ElementTree.parse = defusedxml.ElementTree.parse
xml.etree.ElementTree.iterparse = defusedxml.ElementTree.iterparse

# python_calamine is imported dynamically inside parse_excel_to_json to allow openpyxl fallback

logger = logging.getLogger(__name__)


# =============================================================================
# In-memory LRU cache — keyed by survey_id, max 50 entries
# =============================================================================

_CACHE_MAX_SIZE = 50
_cache: collections.OrderedDict[str, dict] = collections.OrderedDict()
_cache_lock = threading.Lock()


def _cache_get(survey_id: str) -> dict | None:
    """Thread-safe cache lookup. Moves the entry to the end (most recently used)."""
    with _cache_lock:
        if survey_id in _cache:
            _cache.move_to_end(survey_id)
            logger.info(f"Cache HIT for survey {survey_id}")
            return _cache[survey_id]
    return None


def _cache_put(survey_id: str, data: dict) -> None:
    """Thread-safe cache insert. Evicts the oldest entry if over capacity."""
    with _cache_lock:
        _cache[survey_id] = data
        _cache.move_to_end(survey_id)
        while len(_cache) > _CACHE_MAX_SIZE:
            evicted_key, _ = _cache.popitem(last=False)
            logger.info(f"Cache EVICT: {evicted_key}")


def invalidate_cache(survey_id: str) -> None:
    """Remove a survey from the cache (e.g. after deletion)."""
    with _cache_lock:
        if survey_id in _cache:
            del _cache[survey_id]
            logger.info(f"Cache INVALIDATE: {survey_id}")


def compute_file_hash(file_bytes: bytes) -> str:
    """Compute SHA-256 hash of file bytes for deduplication."""
    return hashlib.sha256(file_bytes).hexdigest()


def _to_number(value: Any) -> float | None:
    """Attempt to parse a value as float. Fast-path for int/float."""
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sanitize_data(data: Any) -> Any:
    """Recursively replace invalid JSON floats (NaN, Inf) with None."""
    import math
    if isinstance(data, dict):
        return {k: _sanitize_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [_sanitize_data(v) for v in data]
    elif isinstance(data, float):
        if math.isnan(data) or math.isinf(data):
            return None
    return data


def _extract_titles_from_rows(rows: list[list[Any]]) -> dict[int, str]:
    """Extract table number → title mapping from the INDEX sheet rows."""
    titles: dict[int, str] = {}
    start_row = 3
    for r_idx, row in enumerate(rows, 1):
        if row and len(row) > 0 and row[0] and "Table" in str(row[0]):
            start_row = r_idx + 1
            break

    for row in rows[start_row - 1:]:
        table_cell = row[0] if len(row) > 0 else None
        title = row[1] if len(row) > 1 else None
        if not table_cell or not title:
            continue

        match = re.match(r"Table\s*(\d+)", str(table_cell), re.IGNORECASE)
        if match:
            titles[int(match.group(1))] = str(title).strip()

    return titles


def parse_excel_to_json(
    file_bytes: bytes,
) -> dict[str, Any]:
    """
    Parse an Excel survey workbook entirely in-memory and return structured JSON.
    Uses Calamine as a primary high-performance parser, with an openpyxl fallback
    for environments where Calamine is unavailable.

    Args:
        file_bytes: Raw bytes of the uploaded .xlsx/.xlsm file.

    Returns:
        Dict structured as:
        {
            "table_num": {
                "title": "Question Title",
                "data": {
                    "row_label": { "column_name": numeric_value, ... },
                    ...
                }
            }
        }
    """
    table_titles = {}
    rows = []
    calamine_loaded = False

    # 1. Try Calamine
    try:
        from python_calamine import CalamineWorkbook
        stream = BytesIO(file_bytes)
        try:
            workbook = CalamineWorkbook.from_filelike(stream)
            index_sheet_name = None
            for name in workbook.sheet_names:
                if name.upper() == "INDEX":
                    index_sheet_name = name
                    break

            if index_sheet_name:
                index_rows = workbook.get_sheet_by_name(index_sheet_name).to_python()
                table_titles = _extract_titles_from_rows(index_rows)

            tables_sheet_name = None
            for name in workbook.sheet_names:
                if name.upper() == "TABLES":
                    tables_sheet_name = name
                    break
            if not tables_sheet_name:
                raise ValueError("Missing 'Tables' sheet in the workbook.")

            rows = workbook.get_sheet_by_name(tables_sheet_name).to_python()
            workbook.close()
            calamine_loaded = True
        finally:
            stream.close()
    except Exception as e:
        logger.warning(f"Calamine parser failed or unavailable: {e}. Falling back to openpyxl.")

    # 2. Try openpyxl if Calamine failed
    if not calamine_loaded:
        import openpyxl
        stream = BytesIO(file_bytes)
        try:
            workbook = openpyxl.load_workbook(stream, read_only=True, data_only=True)
            index_sheet_name = None
            for name in workbook.sheetnames:
                if name.upper() == "INDEX":
                    index_sheet_name = name
                    break

            if index_sheet_name:
                index_rows = [list(r) for r in workbook[index_sheet_name].iter_rows(values_only=True)]
                table_titles = _extract_titles_from_rows(index_rows)

            tables_sheet_name = None
            for name in workbook.sheetnames:
                if name.upper() == "TABLES":
                    tables_sheet_name = name
                    break
            if not tables_sheet_name:
                raise ValueError("Missing 'Tables' sheet in the workbook.")

            rows = [list(r) for r in workbook[tables_sheet_name].iter_rows(values_only=True)]
            workbook.close()
        finally:
            stream.close()

    parsed_tables: dict[int, dict] = {}

    # Find all table start positions
    table_starts: list[tuple[str, int]] = []
    for idx, row in enumerate(rows):
        if row and row[0] and str(row[0]).startswith("Table "):
            table_starts.append((row[0], idx))

        for i, (table_num_str, start_idx) in enumerate(table_starts):
            match = re.match(r"Table\s*(\d+)", table_num_str, re.IGNORECASE)
            if not match:
                continue
            table_num = int(match.group(1))

            title = table_titles.get(table_num, "Unknown")
            if title == "Unknown" and start_idx + 1 < len(rows):
                title = str(rows[start_idx + 1][0] or "Unknown").strip()

            # Find the header row containing "TOTAL"
            header_break_idx = None
            for offset in range(1, 20):
                r_idx = start_idx + offset
                if r_idx >= len(rows):
                    break
                if rows[r_idx] and any(
                    isinstance(v, str) and v.strip().upper() == "TOTAL"
                    or (v is not None and not isinstance(v, str) and str(v).strip().upper() == "TOTAL")
                    for v in rows[r_idx]
                ):
                    header_break_idx = r_idx
                    break

            if header_break_idx is None:
                continue

            # Detect two-row headers (parent + sub-column)
            desc_row_idx = header_break_idx - 2
            is_two_row = False
            desc_row: tuple = ()
            if desc_row_idx > start_idx:
                desc_row = rows[desc_row_idx]
                non_none_desc = [
                    v for v in desc_row[1:] if v is not None and str(v).strip() != ""
                ]
                if len(non_none_desc) > 0:
                    is_two_row = True

            break_row = rows[header_break_idx]

            # Build column headers
            if is_two_row:
                column_headers: dict[int, str] = {}
                current_parent = None
                for col_idx in range(1, len(break_row)):
                    sub = break_row[col_idx]
                    parent = desc_row[col_idx] if col_idx < len(desc_row) else None

                    if parent is not None and str(parent).strip() != "":
                        current_parent = str(parent).strip()

                    if sub is None or str(sub).strip() == "":
                        continue

                    sub_str = str(sub).strip()
                    if sub_str.upper() == "TOTAL":
                        col_name = "Total"
                    elif current_parent:
                        col_name = f"{current_parent}: {sub_str}"
                    else:
                        col_name = sub_str

                    column_headers[col_idx] = col_name
                data_start_offset = 11
            else:
                column_headers = {}
                for col_idx, val in enumerate(break_row):
                    if val not in (None, ""):
                        col_name = str(val).strip()
                        column_headers[col_idx] = col_name
                data_start_offset = (header_break_idx - start_idx) + 1

            # Parse data rows
            table_data: dict[str, dict[str, Any]] = {}
            row_idx = start_idx + data_start_offset

            while row_idx < len(rows):
                row = rows[row_idx]
                row_label = str(row[0]).strip() if row and row[0] is not None else ""

                if row_label.startswith("Table ") or row_label == "Sigma":
                    break

                if row_label in ("", " ") and all(v is None for v in row[1:]):
                    row_idx += 1
                    continue

                row_data: dict[str, Any] = {}
                for col_idx, col_name in column_headers.items():
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val not in (None, ""):
                            number = _to_number(val)
                            row_data[col_name] = number if number is not None else val

                if row_data:
                    table_data[row_label] = row_data

                row_idx += 1

            # Extract base rows for two-row header tables
            if is_two_row:
                for offset in [1, 2]:
                    base_row_idx = header_break_idx + offset
                    if base_row_idx < len(rows):
                        base_row = rows[base_row_idx]
                        base_label = str(base_row[0]).strip()
                        if "Base" in base_label or "Sample" in base_label:
                            base_data: dict[str, Any] = {}
                            for col_idx, col_name in column_headers.items():
                                if col_idx < len(base_row):
                                    val = base_row[col_idx]
                                    if val not in (None, ""):
                                        number = _to_number(val)
                                        base_data[col_name] = (
                                            number if number is not None else val
                                        )
                            if base_data:
                                if "unwtd" in base_label.lower() or "unweighted" in base_label.lower():
                                    normalized_label = "Unweighted Sample"
                                elif "wtd" in base_label.lower() or "weighted" in base_label.lower():
                                    normalized_label = "Weighted Sample"
                                else:
                                    normalized_label = (
                                        "Unweighted Sample"
                                        if "un" in base_label.lower()
                                        else "Weighted Sample"
                                    )
                                table_data[normalized_label] = base_data

            parsed_tables[table_num] = table_data

    # Build final output and sanitize NaN/Inf floats
    return _sanitize_data({
        str(table_num): {
            "title": table_titles.get(table_num, "Unknown"),
            "data": table_data,
        }
        for table_num, table_data in sorted(parsed_tables.items())
    })


def get_parsed_survey_data(survey_row: dict, cfg) -> dict:
    """
    Get the fully parsed survey JSON data. Uses an in-memory LRU cache
    keyed by survey_id so that repeated requests (e.g. switching tables,
    computing intersections) don't re-parse the workbook.
    """
    survey_id = survey_row.get("id", "")

    # 1. Check cache first
    cached = _cache_get(survey_id)
    if cached is not None:
        return cached

    survey_data = survey_row.get("survey_data", {})
    if isinstance(survey_data, dict) and "raw_file_b64" in survey_data:
        import base64
        file_bytes = base64.b64decode(survey_data["raw_file_b64"])

        # 4. Parse workbook
        result = parse_excel_to_json(file_bytes)

        # 5. Cache the result
        _cache_put(survey_id, result)

        return result

    return survey_data
